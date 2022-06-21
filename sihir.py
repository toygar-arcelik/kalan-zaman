import os
from xmlrpc.client import MAXINT
import colorama
from openpyxl import Workbook, load_workbook
import pandas as pd
from colorama import Fore, Style, init

vtsPath = r'\\arcrk03V\crk-ortak\Kurutucu\UGY\Sistem_Tasarım\UGY_LABORATUARI\performans_lab_VTS'
vtsLogPath = r'\\arcrk03V\crk-ortak\Kurutucu\UGY\Sistem_Tasarım\UGY_LABORATUARI\performans_lab_VTS\LOG'
perfFilesPath = r'.\PERFORMANS_FILES'
logFilesPath = './LOG/'
wb = Workbook()


perfRowIndex = {
    'TEST NO': 4,
    'KAPASITE': 5,
    'PROGRAM': 9,
    'YUK': 27,
}

summaryParameters = ['MAKINA NO', 'TEST NO', 'RUN ARALIGI', 'ORT ADC', 'ORT DOOR TEMP', 'BLDC RPM', 'BLDC CURRENT', 'KAPASITE',
 'ELAPSED TIME (dk)', 'CAMASIR TIPI', 'WET FINISH TIME', 'WET CONDUCTIVITY COUNTER CUPBOARD', 'WET CONDUCTIVITY COUNTER MIXED',
 'DOOR TEMP AT WET', 'IRON DRY FINISH TIME', 'IRONDRY CONDUCTIVITY COUNTER CUPBOARD', 'IRONDRY CONDUCTIVITY COUNTER MIXED',
 'DOOR TEMP AT IRONDRY', 'CUPBOARD EW TIME', 'DOOR TEMP AT CUPBOARD EW', 'CUPBOARD DRY TIME', 'DOOR TEMP AT CUPBOARD DRY',
 'PROGRAM NO', 'DRYNESS LEVEL', 'BAŞLANGIÇ X', 'BAŞLANGIÇ Y']
excelColumnIndex = dict(zip(summaryParameters, range(1, len(summaryParameters)+1)))


# You can change this variables
timeIntervalCount = 3               
timeIntervals = [
    {'start': 60, 'end': 240},
    {'start': 240, 'end': 500},
    {'start': 500, 'end': 1000},
]

# You cannot change
unilogPrefix = 'UL - '
#selectedProgram, DrynessLevel
def averageValue(df, parameter, startTime, endTime):
    return df[(df['elapsedTime'] >= startTime) & (df['elapsedTime'] <= endTime)][parameter].mean()

def findStartStopIndex(df: pd.DataFrame, column_name, parameter):
    try:
        _idx = df.index[df[column_name] == parameter]
        _startIndex = _idx[0]
        _stopIndex = _idx[-1]
        return _startIndex, _stopIndex
    except:
        return None, None

def findStartStopTime(df: pd.DataFrame, column_name, parameter):
    startIndex, stopIndex = findStartStopIndex(df, column_name, parameter)
    try:
        startTime = df.at[startIndex, 'elapsedTime']
        stopTime = df.at[stopIndex, 'elapsedTime']
        return startTime, stopTime
    except:
        return 0, 0

def findDrynessLevel(df: pd.DataFrame):
    firstRun, lastRun = findStartStopIndex(df, 'state', 'RUN')
    try:
        drynessLevel = df.at[firstRun, 'DrynessLevel']
        return drynessLevel
    except:
        return None

def findSelectedProgram(df: pd.DataFrame):
    firstRun, lastRun = findStartStopIndex(df, 'state', 'RUN')
    try:
        selectedProgram = df.at[firstRun, 'selectedProgram']
        return selectedProgram
    except:
        return None

def findConductivityValues(df: pd.DataFrame, column_name, parameter):
    startIndex, stopIndex = findStartStopIndex(df, column_name, parameter)
    try:
        _conductivityCounterCupboard = df.at[stopIndex, 'conductivityCounterCupboard']
        _conductivityCounterMixed = df.at[stopIndex, 'conductivityCounterMixed']
        return _conductivityCounterCupboard, _conductivityCounterMixed
    except:
        return None, None

def findStartStopDoorTemp(df: pd.DataFrame, column_name, parameter):
    startIndex, stopIndex = findStartStopIndex(df, column_name, parameter)
    try:
        _startDoorTemp = df.at[startIndex, 'doorTemperature']
        _stopDoorTemp = df.at[stopIndex, 'doorTemperature']
        return _startDoorTemp, _stopDoorTemp
    except:
        return None, None

def checkTimeIntervals():
    if timeIntervalCount != len(timeIntervals):
        print('Time interval count is not equal to run time interval count!')
        return False

    for _interval in timeIntervals:
        if _interval['start'] >= _interval['end']:
            print('Start time is greater than end time!')
            return False

    return True

def getMachines():
    machines = []
    for file in os.listdir(perfFilesPath):
        if file.endswith('.xls'):
            machines.append(file.split('_')[0])
    return machines

def updateSummarySheet(_ws, _machineno, _perf_df, _row, _column):
    
    for _keycol in range(1, len(excelColumnIndex) + 1):
        _ws.cell(row=1, column=_keycol).value = list(excelColumnIndex.keys())[_keycol - 1]

    _perf_df_cols = _perf_df.columns.values.tolist()[1:]
    
    _testNo = _perf_df.at[perfRowIndex['TEST NO'], _perf_df_cols[_column]]
    _kapasite = _perf_df.at[perfRowIndex['KAPASITE'], _perf_df_cols[_column]]
    _camasirTipi = _perf_df.at[perfRowIndex['YUK'], _perf_df_cols[_column]]

    if type(_testNo) is not int:
        return

    _ws.cell(row=_row, column=excelColumnIndex['MAKINA NO']).value = _machineno
    _ws.cell(row=_row, column=excelColumnIndex['TEST NO']).value = _testNo
    _ws.cell(row=_row, column=excelColumnIndex['KAPASITE']).value = _kapasite
    _ws.cell(row=_row, column=excelColumnIndex['CAMASIR TIPI']).value = _camasirTipi

    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['MAKINA NO'], end_column=excelColumnIndex['MAKINA NO'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['TEST NO'], end_column=excelColumnIndex['TEST NO'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['KAPASITE'], end_column=excelColumnIndex['KAPASITE'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['CAMASIR TIPI'], end_column=excelColumnIndex['CAMASIR TIPI'])

    # Log incelemenin başladığı yer
    testLogFile = logFilesPath + unilogPrefix + _machineno + '_Test' + str(_testNo) + '.csv'

    if not os.path.exists(testLogFile):
        print(Fore.LIGHTRED_EX + 'Test log file not found: ' + testLogFile)
        return
    else:
        print(Fore.LIGHTGREEN_EX + 'Test log file found: ' + testLogFile)
    
    try:
        # I used low_memory=False to avoid memory error which is worse solution :)
        testLog_df = pd.read_csv(testLogFile, sep=',', skiprows=1, low_memory=False)
    except pd.errors.DtypeWarning:
        print(Fore.LIGHTWHITE_EX + 'Error while reading test log file: ' + testLogFile)

    # Calculations depends on time interval
    for _offset, _interval in zip(range(0, timeIntervalCount), timeIntervals):
        _ws.cell(row=_row+_offset, column=excelColumnIndex['RUN ARALIGI']).value = str(_interval['start']) + '-' + str(_interval['end'])
        _ws.cell(row=_row+_offset, column=excelColumnIndex['ORT DOOR TEMP']).value = averageValue(testLog_df, 'doorTemperature', _interval['start'], _interval['end']).__format__('.2f')
        _ws.cell(row=_row+_offset, column=excelColumnIndex['ORT ADC']).value = averageValue(testLog_df, 'conductivityValue', _interval['start'], _interval['end']).__format__('.2f')
        _ws.cell(row=_row+_offset, column=excelColumnIndex['BLDC RPM']).value = averageValue(testLog_df, 'bldcRPM', _interval['start'], _interval['end']).__format__('.2f')
        _ws.cell(row=_row+_offset, column=excelColumnIndex['BLDC CURRENT']).value = averageValue(testLog_df, 'bldcCurrent', _interval['start'], _interval['end']).__format__('.2f')


    # Advanced calculations
    elapsedTimeStart, elapsedTimeFinish = findStartStopTime(testLog_df, 'state', 'RUN')

    wetStartTime, wetStopTime = findStartStopTime(testLog_df, 'humidityDecision', 'WET')
    irondryStartTime, irondryStopTime = findStartStopTime(testLog_df, 'humidityDecision', 'IRON_DRY')
    cupboardEWStartTime, cupboardEWStopTime = findStartStopTime(testLog_df, 'humidityDecision', 'CUPBOARD_DRY_EW')
    cupboarddryStartTime, cupboarddryStopTime = findStartStopTime(testLog_df, 'humidityDecision', 'CUPBOARD_DRY')

    wetConductivityCounterCupboard, wetConductivityCounterMixed = findConductivityValues(testLog_df, 'humidityDecision', 'WET')
    irondryConductivityCounterCupboard, irondryConductivityCounterMixed = findConductivityValues(testLog_df, 'humidityDecision', 'IRON_DRY')

    wetStartDoorTemp, wetStopDoorTemp = findStartStopDoorTemp(testLog_df, 'humidityDecision', 'WET')
    irondryStartDoorTemp, irondryStopDoorTemp = findStartStopDoorTemp(testLog_df, 'humidityDecision', 'IRON_DRY')
    cupboardEWStartDoorTemp, cupboardEWStopDoorTemp = findStartStopDoorTemp(testLog_df, 'humidityDecision', 'CUPBOARD_DRY_EW')
    cupboarddryStartDoorTemp, cupboarddryStopDoorTemp = findStartStopDoorTemp(testLog_df, 'humidityDecision', 'CUPBOARD_DRY')

    # Assign values to summary sheet
    _ws.cell(row=_row, column=excelColumnIndex['WET FINISH TIME']).value = wetStopTime
    _ws.cell(row=_row, column=excelColumnIndex['WET CONDUCTIVITY COUNTER CUPBOARD']).value = wetConductivityCounterCupboard
    _ws.cell(row=_row, column=excelColumnIndex['WET CONDUCTIVITY COUNTER MIXED']).value = wetConductivityCounterMixed
    _ws.cell(row=_row, column=excelColumnIndex['DOOR TEMP AT WET']).value = wetStopDoorTemp

    _ws.cell(row=_row, column=excelColumnIndex['IRON DRY FINISH TIME']).value = irondryStopTime
    _ws.cell(row=_row, column=excelColumnIndex['IRONDRY CONDUCTIVITY COUNTER CUPBOARD']).value = irondryConductivityCounterCupboard
    _ws.cell(row=_row, column=excelColumnIndex['IRONDRY CONDUCTIVITY COUNTER MIXED']).value = irondryConductivityCounterMixed
    _ws.cell(row=_row, column=excelColumnIndex['DOOR TEMP AT IRONDRY']).value = irondryStopDoorTemp

    _ws.cell(row=_row, column=excelColumnIndex['CUPBOARD EW TIME']).value = cupboardEWStopTime - cupboardEWStartTime
    _ws.cell(row=_row, column=excelColumnIndex['DOOR TEMP AT CUPBOARD EW']).value = cupboardEWStopDoorTemp
    
    _ws.cell(row=_row, column=excelColumnIndex['CUPBOARD DRY TIME']).value = cupboarddryStopTime - cupboarddryStartTime
    _ws.cell(row=_row, column=excelColumnIndex['DOOR TEMP AT CUPBOARD DRY']).value = cupboarddryStopDoorTemp

    _ws.cell(row=_row, column=excelColumnIndex['PROGRAM NO']).value = findSelectedProgram(testLog_df)
    _ws.cell(row=_row, column=excelColumnIndex['DRYNESS LEVEL']).value = findDrynessLevel(testLog_df)
    _ws.cell(row=_row, column=excelColumnIndex['ELAPSED TIME (dk)']).value = round(elapsedTimeFinish/60)


    # Merge Cells
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['WET FINISH TIME'], end_column=excelColumnIndex['WET FINISH TIME'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['WET CONDUCTIVITY COUNTER CUPBOARD'], end_column=excelColumnIndex['WET CONDUCTIVITY COUNTER CUPBOARD'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['WET CONDUCTIVITY COUNTER MIXED'], end_column=excelColumnIndex['WET CONDUCTIVITY COUNTER MIXED'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['DOOR TEMP AT WET'], end_column=excelColumnIndex['DOOR TEMP AT WET'])

    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['IRON DRY FINISH TIME'], end_column=excelColumnIndex['IRON DRY FINISH TIME'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['IRONDRY CONDUCTIVITY COUNTER CUPBOARD'], end_column=excelColumnIndex['IRONDRY CONDUCTIVITY COUNTER CUPBOARD'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['IRONDRY CONDUCTIVITY COUNTER MIXED'], end_column=excelColumnIndex['IRONDRY CONDUCTIVITY COUNTER MIXED'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['DOOR TEMP AT IRONDRY'], end_column=excelColumnIndex['DOOR TEMP AT IRONDRY'])

    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['CUPBOARD EW TIME'], end_column=excelColumnIndex['CUPBOARD EW TIME'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['DOOR TEMP AT CUPBOARD EW'], end_column=excelColumnIndex['DOOR TEMP AT CUPBOARD EW'])

    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['CUPBOARD DRY TIME'], end_column=excelColumnIndex['CUPBOARD DRY TIME'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['DOOR TEMP AT CUPBOARD DRY'], end_column=excelColumnIndex['DOOR TEMP AT CUPBOARD DRY'])

    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['PROGRAM NO'], end_column=excelColumnIndex['PROGRAM NO'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['DRYNESS LEVEL'], end_column=excelColumnIndex['DRYNESS LEVEL'])
    _ws.merge_cells(start_row=_row, end_row=_row+timeIntervalCount-1, start_column=excelColumnIndex['ELAPSED TIME (dk)'], end_column=excelColumnIndex['ELAPSED TIME (dk)'])


def main():
    init(autoreset=True)

    if not checkTimeIntervals():
        return

    machines = getMachines()
    if len(machines) == 0:
        print('No machines found!')
        return

    for perf_file in os.listdir(perfFilesPath):
        if not perf_file.endswith('.xls'):
            continue
        df_perffile = pd.read_excel(os.path.join(perfFilesPath, perf_file), sheet_name=None)
        machineno = perf_file.split('_')[0]
        for sheet in df_perffile.keys():
            if not 'PERF' in sheet:
                continue
            perf_df = pd.read_excel(perfFilesPath + '\\' + perf_file, sheet_name=sheet)
            ws = wb.create_sheet(machineno+''+sheet, index=len(wb.sheetnames))
            for _row, _column in zip(range(2, MAXINT, timeIntervalCount),
                                    range(0, perf_df.columns.__len__()-1)):
                updateSummarySheet(ws, machineno, perf_df, _row, _column)
        

    wb.save('genel_ozet.xlsx')
    wb.close()

if __name__ == '__main__':
    main()